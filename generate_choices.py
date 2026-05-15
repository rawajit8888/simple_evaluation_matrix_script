import unicodedata
import openpyxl
import os

def clean(text):
    """Remove hidden unicode characters and normalize."""
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFC", text)
    text = text.replace('\u00a0', ' ')   # non-breaking space
    text = text.replace('\u200b', '')    # zero-width space
    text = text.replace('\u200c', '')    # zero-width non-joiner
    text = text.replace('\u200d', '')    # zero-width joiner
    text = text.replace('\ufeff', '')    # BOM
    text = text.replace('\uff1e', '>')   # fullwidth >
    text = ' '.join(text.split())
    return text.strip()

def to_choice(text):
    """Convert to Label Studio <Choice> XML line."""
    text = text.replace('&', '&amp;')
    text = text.replace('>', '&gt;')
    text = text.replace('<', '&lt;')
    return f'<Choice value="{text}" />'

# -------------------------------------------------------
# CHANGE THIS to your actual xlsx filename
INPUT_FILE = "testrun.xlsx"
# -------------------------------------------------------

wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
ws = wb.active

# Read headers
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i for i, name in enumerate(headers)}

print(f"Reading: {INPUT_FILE}")
print(f"Columns: {headers}\n")

# Extract category_desc > categ_type_desc > category_subtype_desc
seen = set()
choices = []

for row in ws.iter_rows(min_row=2, values_only=True):
    cat1 = clean(row[col['category_desc']])
    cat2 = clean(row[col['categ_type_desc']])
    cat3 = clean(row[col['category_subtype_desc']])

    if cat1 and cat2 and cat3:
        combined = f"{cat1} > {cat2} > {cat3}"
        if combined not in seen:
            seen.add(combined)
            choices.append(combined)

print(f"Found {len(choices)} unique categories\n")

# Save output XML
output_file = INPUT_FILE.replace(".xlsx", "_choices.xml")
with open(output_file, "w", encoding="utf-8") as f:
    for c in choices:
        line = to_choice(c)
        f.write(line + "\n")
        print(line)

print(f"\nDone! Saved to: {output_file}")
